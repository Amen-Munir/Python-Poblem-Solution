# automatng xlsx calculations of any number of products whose prices needs to be re-calculated 
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb=xl.load_workbook('demo1.xlsx')
sheet=wb['Sheet1']
cell=sheet.cell(1,1)

for row in range(2, sheet.max_row+1):
    cell=sheet.cell(row,3)
    new_price=cell.value*0.8
    new_price_cell=sheet.cell(row,4)
    new_price_cell.value=new_price

val=Reference(sheet, min_row= 2, max_row=sheet.max_row, min_col=4, max_col=4)
chart=BarChart()
chart.add_data(val)
sheet.add_chart(chart, 'f1')
wb.save('demo2.xlsx')
